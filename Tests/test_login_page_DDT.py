"""Testing login with Data driven approach via excell"""
import os

import openpyxl
import pytest

from Configuration.CostumeLogger import CostumeLogger
from Configuration.TestData import TestData
from Pages.LoginPage import LoginPage
from Tests.TestBase import BaseTest


class TestLoginPage(BaseTest):
    log = CostumeLogger.log()

    @pytest.mark.ddt_login
    def test_login_ddt_with_excel(self):
        self.log.info('--------------- Testing login ddt excel ---------------')
        file_path = os.path.join(os.getcwd(), 'test-data', 'login_data_example.xlsx')
        wb = openpyxl.load_workbook(file_path)
        sheet = wb['login_data']
        num_rows = sheet.max_row
        self.loginPage = LoginPage(self.driver)
        for i in range(2, num_rows + 1):

            self.loginPage.login(sheet.cell(i, 1).value, sheet.cell(i, 2).value)
            self.log.info('Username: ' + sheet.cell(i, 1).value)
            self.log.info('Password: ' + sheet.cell(i, 2).value)
            self.log.info('Expected results: ' + str(sheet.cell(i, 3).value))
            if self.driver.current_url == TestData.DASHBOARD_URL:
                login_success = True
                self.log.info('--------------- Login was successful ---------------')
                self.loginPage.logout()
            else:
                login_success = False
                self.log.info('--------------- Login was not successful ---------------')

            assert login_success == sheet.cell(i, 3).value, 'Login was not successful!'


