import os

import openpyxl
import pytest

from Configuration.CostumeLogger import CostumeLogger
from Pages.PersonalInfoPage import PersonalPageInfo
from Tests.TestBase import BaseTest


class TestMyPersonalInfoPageDDT(BaseTest):
    log = CostumeLogger.log()

    @pytest.mark.update_employee_info_ddt
    def test_update_employee_info(self):
        self.log.info('--------------- Update employee info testing ddt---------------')
        file_path = os.path.join(os.getcwd(),'test-data', 'login_data_example.xlsx')
        wb = openpyxl.load_workbook(file_path)
        sheet = wb['personal_info_data']
        num_rows = sheet.max_row

        self.myinfopage = PersonalPageInfo(self.driver)

        for i in range(2, num_rows+1):

            self.myinfopage.set_first_name(sheet.cell(i,1).value)
            self.myinfopage.set_middle_name(sheet.cell(i,2).value)
            self.myinfopage.set_last_name(sheet.cell(i,3).value)
            self.myinfopage.set_employee_id(str(sheet.cell(i,4).value))
            self.myinfopage.set_other_id(str(sheet.cell(i,5).value))
            self.myinfopage.set_drivers_license_number(str(sheet.cell(i,6).value))
            self.myinfopage.set_license_expiration_date(sheet.cell(i,7).value.strip("'"))
            self.myinfopage.set_nationality(sheet.cell(i,8).value)
            self.myinfopage.set_marital_status(sheet.cell(i,9).value)
            self.myinfopage.set_birth_date(sheet.cell(i,10).value.strip("'"))
            self.myinfopage.set_gender(sheet.cell(i,11).value)
            notification = self.myinfopage.save_changes()

        assert 'Successfully Updated' in notification.text, "Changes are not saved"